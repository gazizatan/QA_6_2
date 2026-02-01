import json
import os
import time
from pathlib import Path

import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait

ROOT = Path(__file__).resolve().parents[1]
DATA_FILE = ROOT / "data" / "test_cases.xlsx"
ARTIFACTS_DIR = ROOT / "artifacts"
SCREENSHOT_DIR = ARTIFACTS_DIR / "screenshots"
SESSIONS_FILE = ARTIFACTS_DIR / "sauce_sessions.json"
RESULTS_FILE = ARTIFACTS_DIR / "test_results.json"

if RESULTS_FILE.exists():
    RESULTS_FILE.unlink()


def load_cases():
    if not DATA_FILE.exists():
        raise FileNotFoundError(f"Missing test data file: {DATA_FILE}")
    wb = load_workbook(DATA_FILE)
    ws = wb.active
    cases = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        case_id, name, url, title_contains, text_contains, case_type = row
        cases.append(
            {
                "id": int(case_id),
                "name": str(name),
                "url": str(url),
                "title_contains": (title_contains or "").strip(),
                "text_contains": (text_contains or "").strip(),
                "type": (case_type or "").strip(),
            }
        )
    return cases


CASES = load_cases()
BROWSERS = ["chrome", "firefox"]


def _get_credentials():
    username = os.getenv("SAUCE_USERNAME")
    access_key = os.getenv("SAUCE_ACCESS_KEY")
    if not username or not access_key:
        raise RuntimeError(
            "Missing Sauce credentials. Set SAUCE_USERNAME and SAUCE_ACCESS_KEY."
        )
    return username, access_key


def _resolve_remote_url():
    remote_url = os.getenv("SAUCE_REMOTE_URL")
    if remote_url:
        return remote_url

    region = os.getenv("SAUCE_REGION", "eu-central-1")
    return f"https://ondemand.{region}.saucelabs.com/wd/hub"


def _create_driver(browser_name: str, test_name: str):
    username, access_key = _get_credentials()

    if browser_name == "chrome":
        options = webdriver.ChromeOptions()
    elif browser_name == "firefox":
        options = webdriver.FirefoxOptions()
    else:
        raise ValueError(f"Unsupported browser: {browser_name}")

    options.browser_version = os.getenv("BROWSER_VERSION", "latest")
    options.platform_name = os.getenv("PLATFORM_NAME", "Windows 11")

    build_name = os.getenv("SAUCE_BUILD", f"QA6-2-{time.strftime('%Y%m%d-%H%M%S')}")

    sauce_options = {
        "username": username,
        "accessKey": access_key,
        "build": build_name,
        "name": test_name,
        "screenResolution": "1920x1080",
        "seleniumVersion": "4.16.0",
        "recordVideo": True,
        "recordScreenshots": True,
        "extendedDebugging": True,
    }

    options.set_capability("sauce:options", sauce_options)

    remote_url = _resolve_remote_url()
    return webdriver.Remote(command_executor=remote_url, options=options)


def _append_session_info(session_info):
    ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)
    if SESSIONS_FILE.exists():
        try:
            existing = json.loads(SESSIONS_FILE.read_text())
        except json.JSONDecodeError:
            existing = []
    else:
        existing = []
    existing.append(session_info)
    SESSIONS_FILE.write_text(json.dumps(existing, indent=2))


def _append_result(result):
    ARTIFACTS_DIR.mkdir(parents=True, exist_ok=True)
    if RESULTS_FILE.exists():
        try:
            existing = json.loads(RESULTS_FILE.read_text())
        except json.JSONDecodeError:
            existing = []
    else:
        existing = []
    existing.append(result)
    RESULTS_FILE.write_text(json.dumps(existing, indent=2))


def _build_steps(case):
    parts = [f"Open {case['url']}", "Wait for page to load"]
    if case["title_contains"]:
        parts.append(f"Verify title contains '{case['title_contains']}'")
    if case["text_contains"]:
        parts.append(f"Verify page contains '{case['text_contains']}'")
    return "; ".join(parts)


def _build_expected(case):
    parts = []
    if case["title_contains"]:
        parts.append(f"Title contains '{case['title_contains']}'")
    if case["text_contains"]:
        parts.append(f"Page contains '{case['text_contains']}'")
    return "; ".join(parts) if parts else "Page loads successfully"


@pytest.mark.parametrize("browser", BROWSERS)
@pytest.mark.parametrize("case", CASES, ids=[c["name"] for c in CASES])
def test_browserstack_site(browser, case):
    test_name = f"{case['id']} - {case['name']} [{browser}]"
    driver = _create_driver(browser, test_name)
    session_id = driver.session_id
    error = None
    passed = False
    actual_title = ""
    text_found = None

    try:
        driver.get(case["url"])
        WebDriverWait(driver, 30).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

        actual_title = driver.title

        if case["title_contains"]:
            assert case["title_contains"] in actual_title

        if case["text_contains"]:
            source_lower = driver.page_source.lower()
            text_found = case["text_contains"].lower() in source_lower
            assert text_found

        passed = True
    except Exception as exc:  # noqa: BLE001
        error = exc
    finally:
        SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
        screenshot_path = SCREENSHOT_DIR / f"{case['id']}_{browser}.png"
        try:
            driver.get_screenshot_as_file(str(screenshot_path))
        except Exception:
            pass

        try:
            driver.execute_script(
                f"sauce:job-result={'passed' if passed else 'failed'}"
            )
        except Exception:
            pass

        _append_session_info(
            {
                "session_id": session_id,
                "browser": browser,
                "test_name": test_name,
                "passed": passed,
            }
        )

        expected = _build_expected(case)
        actual_parts = []
        if actual_title:
            actual_parts.append(f"Title: {actual_title}")
        if text_found is not None:
            actual_parts.append(
                f"Text '{case['text_contains']}' found: {str(text_found)}"
            )
        actual_result = "; ".join(actual_parts) if actual_parts else "N/A"

        _append_result(
            {
                "id": case["id"],
                "name": case["name"],
                "browser": browser,
                "url": case["url"],
                "steps": _build_steps(case),
                "expected_result": expected,
                "actual_result": actual_result,
                "status": "Passed" if passed else "Failed",
                "comments": "" if passed else str(error),
                "session_id": session_id,
                "screenshot": str(screenshot_path),
            }
        )

        try:
            driver.quit()
        except Exception:
            pass

    if error:
        raise error
